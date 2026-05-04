"""Exportação de relatórios para Excel (.xlsx) com formatação."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
def _excel_col_width_from_text(max_chars: float, minimum: float = 14, maximum: float = 72) -> float:
    """Largura aproximada no Excel (unidades ~ largura de caractere '0')."""
    return max(minimum, min(maximum, max_chars * 1.12 + 3.5))


def build_relatorio_distribuicao_xlsx(
    *,
    periodo_titulo: str,
    tipo_label: str,
    col_ref: str,
    col_qtd: str,
    labels: list,
    values: list,
    total: int,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'

    title_fill = PatternFill(fill_type='solid', fgColor='FF1E3A5F')
    meta_fill = PatternFill(fill_type='solid', fgColor='FF2563EB')
    header_fill = PatternFill(fill_type='solid', fgColor='FF475569')
    font_title = Font(bold=True, size=14, color='FFFFFF')
    font_meta_label = Font(bold=True, size=11, color='FFFFFF')
    font_meta_value = Font(size=11, color='FFFFFF')
    font_header = Font(bold=True, size=11, color='FFFFFF')
    font_data = Font(size=11, color='0F172A')
    font_total_label = Font(bold=True, size=11, color='0F172A')
    font_total_value = Font(bold=True, size=12, color='1E3A5F')

    align_wrap = Alignment(vertical='top', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_right_num = Alignment(horizontal='right', vertical='top')

    r = 1
    ws.merge_cells('A1:B1')
    c_title = ws.cell(r, 1, 'SIGAD — relatório de distribuições')
    c_title.font = font_title
    c_title.fill = title_fill
    c_title.alignment = align_center
    r += 1

    ws.cell(r, 1, 'Período').font = font_meta_label
    ws.cell(r, 1).fill = meta_fill
    ws.cell(r, 1).alignment = align_left
    ws.cell(r, 2, periodo_titulo).font = font_meta_value
    ws.cell(r, 2).fill = meta_fill
    ws.cell(r, 2).alignment = align_wrap
    r += 1

    ws.cell(r, 1, 'Tipo de relatório').font = font_meta_label
    ws.cell(r, 1).fill = meta_fill
    ws.cell(r, 1).alignment = align_left
    ws.cell(r, 2, tipo_label).font = font_meta_value
    ws.cell(r, 2).fill = meta_fill
    ws.cell(r, 2).alignment = align_wrap
    r += 1

    r += 1

    ws.cell(r, 1, col_ref).font = font_header
    ws.cell(r, 1).fill = header_fill
    ws.cell(r, 1).alignment = align_left
    ws.cell(r, 2, col_qtd).font = font_header
    ws.cell(r, 2).fill = header_fill
    ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_row = r
    r += 1

    first_data_row = r
    for lab, val in zip(labels, values):
        ws.cell(r, 1, lab).font = font_data
        ws.cell(r, 1).alignment = align_wrap
        ws.cell(r, 2, val).font = font_data
        ws.cell(r, 2).alignment = align_right_num
        r += 1

    r += 1
    ws.cell(r, 1, 'Total no período').font = font_total_label
    ws.cell(r, 1).alignment = align_left
    ws.cell(r, 2, total).font = font_total_value
    ws.cell(r, 2).alignment = align_right_num

    max_len_a = max(
        len(str(col_ref)),
        len('SIGAD — relatório de distribuições') // 2,
        max((len(str(x)) for x in labels), default=0),
        len('Total no período'),
    )
    max_len_b = max(len(str(col_qtd)), len(str(periodo_titulo)), len(str(tipo_label)), len(str(total)))

    ws.column_dimensions['A'].width = _excel_col_width_from_text(max_len_a, minimum=28, maximum=78)
    ws.column_dimensions['B'].width = _excel_col_width_from_text(max_len_b, minimum=22, maximum=48)

    ws.row_dimensions[1].height = 28
    for row_idx in range(2, 4):
        ws.row_dimensions[row_idx].height = 22
    ws.row_dimensions[header_row].height = 22
    for row_idx in range(first_data_row, first_data_row + len(labels)):
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = ws.cell(first_data_row, 1)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
